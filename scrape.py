import win32com.client
import os
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, messagebox

# Path to the Excel file
EXCEL_FILE = "Shift_Reports.xlsx"

# Predefined filters
PREDEFINED_FILTERS = {
    "Shift Report": ["Shift Report", "Maintenance shift report", "mech shift report"],
    "Maintenance Only": ["Maintenance shift report"],
    "Custom Filter": []  # To be customized by the user
}

# Lines to remove
LINES_TO_REMOVE = ["*", "<", "Confidentiality Warning", "AUTOMATION TECHNCIAN", "Joseph Dengler"]

def clean_body(body):
    """Clean the email body by removing unwanted lines."""
    lines = body.splitlines()
    cleaned_lines = [
        line.strip() for line in lines 
        if line.strip() and not any(remove in line for remove in LINES_TO_REMOVE)
    ]
    return "\n".join(cleaned_lines)

def adjust_excel_formatting(sheet):
    """Set cell formatting to wrap text, adjust row heights, and sort by date."""
    # Auto-adjust cell formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        # Adjust the row height based on the longest text in the row
        max_length = max(len(str(cell.value) if cell.value else "") for cell in row)
        sheet.row_dimensions[row[0].row].height = min(15 + (max_length // 40) * 15, 300)
    
    # Sort the sheet by Received Time (Column C), descending order (newest first)
    rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True))
    rows_sorted = sorted(rows, key=lambda x: x[2], reverse=True)  # Sort by date (column 3)
    
    # Clear existing rows and write the sorted rows
    for row_index, row in enumerate(rows_sorted, start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

def scrape_outlook(filter_keywords):
    try:
        # Connect to Outlook
        print("Connecting to Outlook...")
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

        # Access the root folder (adjust as needed)
        root_folder = outlook.Folders("Joseph.Dengler@bridor.com")
        inbox = root_folder.Folders("Inbox")

        print(f"Accessed folder: {inbox.Name}")
        messages = inbox.Items

        # Remove the date restriction to process all emails
        print("Fetching all emails...")

        # Load or create the Excel workbook
        if os.path.exists(EXCEL_FILE):
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            print(f"Loaded existing Excel file: {EXCEL_FILE}")
        else:
            workbook = openpyxl.Workbook()
            print(f"Created new Excel workbook: {EXCEL_FILE}")

        sheet = workbook.active
        if sheet.max_row == 1:  # Add headers if the sheet is empty
            sheet.append(["Subject", "Sender", "Received Time", "Body"])

        # Collect existing data to prevent duplicates
        existing_data = {
            (sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=3).value)
            for i in range(2, sheet.max_row + 1)
        }

        # Process emails and populate the Excel file
        for message in messages:
            try:
                subject = message.Subject
                sender = message.SenderName
                received_time = message.ReceivedTime.strftime("%Y-%m-%d %H:%M:%S")
                body = clean_body(message.Body)  # Clean the body of the email

                # Filter emails by keywords
                if not any(keyword.lower() in subject.lower() for keyword in filter_keywords):
                    continue

                # Skip duplicates
                if (subject, received_time) in existing_data:
                    print(f"Skipping duplicate email: {subject}")
                    continue

                # Append email data to the Excel sheet
                sheet.append([subject, sender, received_time, body])
                print(f"Added email: {subject}")

            except Exception as e:
                print(f"Error processing email: {e}")

        # Adjust Excel formatting (wrap text and sort by date)
        adjust_excel_formatting(sheet)

        # Save the updated Excel workbook
        workbook.save(EXCEL_FILE)
        print(f"Excel file updated: {EXCEL_FILE}")

    except Exception as e:
        print(f"Error accessing Outlook: {e}")

def run_gui():
    """Run the GUI for filter selection."""
    def start_script():
        selected_filter = filter_var.get()
        if selected_filter == "Custom Filter":
            custom_filter = custom_filter_entry.get()
            if not custom_filter:
                messagebox.showerror("Error", "Please enter custom filter keywords.")
                return
            filter_keywords = [keyword.strip() for keyword in custom_filter.split(",")]
        else:
            filter_keywords = PREDEFINED_FILTERS[selected_filter]

        # Run the main script with the selected filters
        scrape_outlook(filter_keywords)

    # Initialize GUI window
    root = tk.Tk()
    root.title("Shift Reports Filter Selection")

    # Dropdown for filter selection
    tk.Label(root, text="Select Filter:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
    filter_var = tk.StringVar(value=list(PREDEFINED_FILTERS.keys())[0])
    filter_dropdown = ttk.Combobox(root, textvariable=filter_var, values=list(PREDEFINED_FILTERS.keys()))
    filter_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    # Custom filter entry
    tk.Label(root, text="Custom Filter Keywords (comma-separated):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
    custom_filter_entry = tk.Entry(root, width=40)
    custom_filter_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

    # Start button
    start_button = tk.Button(root, text="Run Script", command=start_script)
    start_button.grid(row=2, column=0, columnspan=2, pady=10)

    root.mainloop()

if __name__ == "__main__":
    run_gui()
